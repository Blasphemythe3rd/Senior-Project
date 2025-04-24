import os
from django.shortcuts import render, redirect
from django.http import Http404, HttpResponse, JsonResponse, HttpResponseNotFound
from django.views.decorators.http import require_GET, require_POST
from PPST.models import Doctor, Test, Stimuli_Response, Given_Stimuli, Notification
from django.core.mail import send_mail
from django.conf import settings
from django.http import FileResponse
from django.contrib import messages
from django.contrib.auth import authenticate, login, get_user, logout
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from django.utils.crypto import get_random_string
import tempfile
import logging
import json
import csv
import tempfile
import json
import pandas as pd
import random
from datetime import datetime
from django.db.models import Avg
from openpyxl import Workbook, load_workbook # type: ignore

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO)

@csrf_exempt  # Use csrf_exempt if you're not using CSRF tokens, though it's recommended to use CSRF tokens in production
def save_response(request):
    if request.method == 'POST':
        # Parse the incoming JSON data from the request
        data = json.loads(request.body)

        # Extract variables from the request data
        test_id = data.get('test_id')
        given_stimuli = data.get('given_stimuli')
        response = data.get('response')
        response_per_click = data.get('response_per_click')
        enum_type = data.get('enum_type')

        # Retrieve the related 'Test' object using the test_id
        test = Test.objects.get(test_id=test_id)

        # Retrieve the related 'Given_Stimuli' object (you can adjust the logic here to match how you want to match stimuli)
        given = Given_Stimuli.objects.get(given_stimuli=given_stimuli)

        # Save the response data to the Stimuli_Response model
        stimuli_response = Stimuli_Response(
            test=test,
            given=given,
            response=response,
            response_per_click=response_per_click,
            enum_type=enum_type
        )
        stimuli_response.save()

        return JsonResponse({"message": "Response saved successfully."}, status=200)

    return JsonResponse({"message": "Invalid request method."}, status=400)


# Store tokens temporarily (use a database model for production)
reset_tokens = {}

def practiceTest(request):
    return render(request,'practiceTest.html')

# def testScreen(request):
#     stimuli_objects = Given_Stimuli.objects.all()
#     stimuli_list = [stimulus.given_stimuli for stimulus in stimuli_objects]

#     return render(request, 'testScreen.html', {'stimuli_list': stimuli_list})

def doctorHomePage(request, username):
    doctors = Doctor.objects.first()
    selected_doctor_id = request.GET.get('doctor')

    try:
        # Fetch the doctor by username
        selected_doctor = Doctor.objects.get(username=username)
        # Fetch notifications for the selected doctor
        notifications = Notification.objects.filter(users=selected_doctor)
    except Doctor.DoesNotExist:
        # If the doctor does not exist, raise a 404 error
        raise Http404("Doctor not found")

    return render(request, 'doctorHomePage.html', {
        'doctors': doctors,
        'notifications': notifications,
        'selected_doctor_id': selected_doctor_id,
        'doctor': selected_doctor  # Pass the doctor object here
    })

def testScreen(request, testId):
    # Check if the testId exists in the database
    test_exists = Test.objects.filter(test_id=testId).exists()

    if not test_exists:
        return HttpResponseNotFound("Test ID not found. Please contact your doctor.")  # Return a 404 response if test_id is invalid

    # Fetch the valid test object
    test_instance = Test.objects.get(test_id=testId)

    # Retrieve all stimuli objects
    stimuli_objects = Given_Stimuli.objects.all()
    stimuli_list = [stimulus.given_stimuli for stimulus in stimuli_objects]
    stimuli_enum = [stimulus.enum_type for stimulus in stimuli_objects]

    # gs = list(Given_Stimuli.objects.all())
    # stimuli = gs[0]

    return render(request, 'testScreen.html', {
        'stimuli_list': stimuli_list,
        'test_id': test_instance.test_id,
        'stimuli_objects': stimuli_objects,
        'stimuli_enum': stimuli_enum
    })

def test(request):
    return HttpResponse("Hello World!")
  
@require_POST
def createTest(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body) #parse the json data from the request
            recipient_email = data.get("email")
            patient_age = data.get("age")
            language = data.get("language")

            if not recipient_email: #check for email
                return JsonResponse({"error": "No email provided"}, status=400)

            doctor = Doctor.objects.first() #get the first doctor (hardcoded for now will change when I can)
            if not doctor:
                return JsonResponse({"error": "No doctor available in the system"}, status=500)

            new_test = Test.objects.create(patient_age=patient_age, doctor=doctor) #create a new test object
            
            
            subject = "Test Link for PPST"
            message = f"A new test has been created for you with ID: http://127.0.0.1:8000/PPST/settings/{new_test.test_id}/{language}" #message to be sent to the user with link(link will need to be changed)
            
            send_mail(subject, message, settings.EMAIL_HOST_USER, [recipient_email])

            return JsonResponse({"message": "Email sent successfully!",
                                 "test_id": new_test.test_id,
                                 "doctor": doctor.last_name,
                                })  
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=500)  

@require_GET
def ppstCreate(request): #this shows the site
    return render(request, "createTest.html", {})

def generateTest(request): #this does nothing right now its just so link kinda works but not yet i need other peoples implementations
    return render(request, "generateTest.html", {})

def doctor_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('PPST:doctorHomePage')
        else:
            # Invalid login
            messages.error(request, 'Invalid username or password')
            return render(request, 'doctorLoginInitial.html')  # Ensure a response is returned
    else:
        return render(request, 'doctorLoginInitial.html')

def logout_view(request):
    logout(request)
    return redirect('PPST:doctor_login')

def testInfo(request):
    percentages = []
    notEmpty = False

    currDoc = get_user(request)
    #currDoc = Doctor.objects.get(username = "doctor0") # use for debugging
    tests = Test.objects.filter(doctor = currDoc)

    for test in tests: # calculate correct response %
        percentages.append(_calculate_accuracies(test))

    zippedTests = zip(tests, percentages) # can use both in django for each:  {% for test, percentage in tests %}

    if(tests): # allows webpage to display no tests screen rather than empty table
        notEmpty = True

    return render(request, "testInfo.html", {
        'doctor' : currDoc,
        'notEmpty' : notEmpty,
        'tests' : zippedTests
    })

def download_test(request, test_id):
    test = Test.objects.get(test_id = test_id)
    
    # uses django's os to grab the template independent of OS
    template_file_path = os.path.join(
        settings.BASE_DIR,  # base directory is the root of the project
        'PPST', 'test_data_template.xlsx' # base_dir/PPST/test_data_template.xlsx
    )

    wb = load_workbook(filename= template_file_path)

    ws = wb["Raw Data"]

    test_info = [test.test_id, test.time_started.replace(tzinfo=None), test.time_ended.replace(tzinfo=None), test.status, test.patient_age]

    cell_range = ws['A2': 'E2']
    index = 0
    for row in cell_range:
        for cell in row:
            cell.value = test_info[index]
            index += 1
        
    responses = Stimuli_Response.objects.filter(test = test)
    response_data = []
    stimulus_num = 1
    latencies = []
    for response in responses:
        response_list = [stimulus_num, response.given.correct_order, response.response, str(response.response_per_click), response.enum_type]
        response_data.append(response_list)
        latencies.append(response.response_per_click)
        stimulus_num += 1

    for row in response_data:
        ws.append(row)

    ws2 = wb["Data Insights"]
    
    index = 0
    for row in ws2.iter_cols(min_row=2, min_col = 4, max_col = 4, max_row = 15):
        for cell in row:
            cell.value = sum(latencies[index]) / len(latencies[index])
            index += 1

    index = 0
    for row in ws2.iter_cols(min_row=2, min_col = 3, max_col = 3, max_row = 15):
        for cell in row:
            cell.value = sum(latencies[index])
            index += 1


    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        tmpfile_path = tmp.name
    
    response = FileResponse(open(tmpfile_path, 'rb'), as_attachment=True, filename="test_data.xlsx")

    return response

def _calculate_accuracies(test):
    QUESTIONS_PER_TEST = 14

    responses = Stimuli_Response.objects.filter(test = test)
    correct = 0
    for response in responses:
        if(response.response == response.given.correct_order):
            correct+=1

    return round(correct/QUESTIONS_PER_TEST * 100, 2)

@csrf_exempt
def admin_login(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            username = data.get('username')
            password = data.get('password')

            logger.info(f"Attempting to authenticate user: {username}")

            user = authenticate(request, username=username, password=password)

            if user is not None and user.is_staff:  
                login(request, user)  
                logger.info(f"User {username} authenticated successfully")
                return JsonResponse({'success': True})
            else:
                logger.warning(f"Authentication failed for user: {username}")
                return JsonResponse({'success': False, 'error': 'Invalid credentials or not an admin.'})
        except Exception as e:
            logger.error(f"Error during login: {e}")
            return JsonResponse({'success': False, 'error': 'An error occurred. Please try again.'})
    else:
        return JsonResponse({'success': False, 'error': 'Invalid request method.'})


def admin_dashboard(request):
    percentages = []
    notEmpty = False

    tests = Test.objects.all()

    for test in tests: # calculate correct response %
        percentages.append(_calculate_accuracies(test))

    zippedTests = zip(tests, percentages) # can use both in django for each:  {% for test, percentage in tests %}

    if(tests): # allows webpage to display no tests screen rather than empty table
        notEmpty = True

    return render(request, "testInfo.html", {
        'notEmpty' : notEmpty,
        'tests' : zippedTests
    })

def admin_page(request):
     try:
         # Fetch all tests
         tests = Test.objects.all()
 
         # Default values if no tests exist
         total_tests = tests.count()
         average_patient_age = tests.aggregate(Avg('patient_age'))['patient_age__avg'] or 0
         total_accuracy = 0
         total_responses = 0
 
         # Calculate overall accuracy
         for test in tests:
             stimuli_responses = Stimuli_Response.objects.filter(test=test)
             for response in stimuli_responses:
                 if response.response and response.given and response.given.given_stimuli:
                     total_responses += 1
                     if str(response.response) == str(response.given.given_stimuli):  # Ensure both are strings for comparison
                         total_accuracy += 1
 
         average_accuracy = (total_accuracy / total_responses * 100) if total_responses > 0 else 0
 
         # Fetch age distribution and accuracy data
         age_data = { "0-9": 0, "10-19": 0, "20-29": 0, "30-39": 0, "40-49": 0,
                      "50-59": 0, "60-69": 0, "70-79": 0, "80-89": 0, "90+": 0 }
         accuracy_data = {key: 0 for key in age_data.keys()}  # Initialize accuracy as 0
 
         if tests.exists():
             # Convert to DataFrame
             df = pd.DataFrame({'patient_age': list(tests.values_list('patient_age', flat=True))})
 
             # Define age bins and labels
             bins = [0, 9, 19, 29, 39, 49, 59, 69, 79, 89, float('inf')]
             labels = ["0-9", "10-19", "20-29", "30-39", "40-49", 
                       "50-59", "60-69", "70-79", "80-89", "90+"]
 
             # Categorize ages into bins
             df['age_group'] = pd.cut(df['patient_age'], bins=bins, labels=labels, right=True)
 
             # Count occurrences in each group
             age_data = df['age_group'].value_counts().sort_index().to_dict()
 
             # Initialize accuracy tracking
             accuracy_data = {key: [] for key in labels}  
 
             # Fetch all responses and match them to the correct answers
             responses = Stimuli_Response.objects.select_related('given')
 
             for response in responses:
                 correct = response.given.correct_order.strip()
                 user_response = response.response.strip()
 
                 if len(correct) == len(user_response) and len(correct) > 0:
                     # Calculate exact character match percentage
                     match_count = sum(1 for c1, c2 in zip(correct, user_response) if c1 == c2)
                     accuracy_percentage = (match_count / len(correct)) * 100
                 else:
                     accuracy_percentage = 0  # If lengths don't match, assume 0% accuracy
 
                 # Get the corresponding test age and categorize it
                 age = response.test.patient_age
                 age_group = pd.cut([age], bins=bins, labels=labels, right=True)[0]
 
                 if age_group in accuracy_data:
                     accuracy_data[age_group].append(accuracy_percentage)
 
             # Compute average accuracy for each age group
             for key in accuracy_data.keys():
                 if accuracy_data[key]:  # Avoid division by zero
                     accuracy_data[key] = sum(accuracy_data[key]) / len(accuracy_data[key])
                 else:
                     accuracy_data[key] = 0  # If no data, default to 0%
 
         # Pass data to the template
         doctors = Doctor.objects.all()
         return render(request, "admin.html", {
             "doctors": doctors,
             "total_tests": total_tests,
             "average_patient_age": average_patient_age,
             "average_accuracy": f"{average_accuracy:.2f}",
             "age_labels": json.dumps(list(age_data.keys())),
             "age_values": json.dumps(list(age_data.values())),
             "accuracy_labels": json.dumps(list(accuracy_data.keys())),
             "accuracy_values": json.dumps(list(accuracy_data.values())),
         })
 
     except Exception as e:
         return render(request, "admin.html", {
             "error": f"An error occurred: {str(e)}",
             "doctors": [],
             "total_tests": 0,
             "average_patient_age": 0,
             "average_accuracy": 0,
             "age_labels": json.dumps([]),
             "age_values": json.dumps([]),
             "accuracy_labels": json.dumps([]),
             "accuracy_values": json.dumps([]),
         })

@require_GET
def list_doctors(request):
    doctors = Doctor.objects.all().values("id", "first_name", "last_name", "email", "username")
    return JsonResponse(list(doctors), safe=False)


@require_GET
def fetch_all_tests(request):
    try:
        # Fetch all tests and their associated doctors
        tests = Test.objects.select_related('doctor').all()
        test_data = []

        for test in tests:
            # Dynamically calculate the correct percentage
            correct_percentage = _calculate_accuracies(test)  # Replace with your actual calculation function

            test_data.append({
                "test_id": test.test_id,
                "time_started": test.time_started.strftime('%d/%m/%y %H:%M') if test.time_started else None,
                "time_ended": test.time_ended.strftime('%d/%m/%y %H:%M') if test.time_ended else None,
                "status": "Not Started" if test.status == 0 else "In Progress" if test.status == 1 else "Complete",
                "patient_age": test.patient_age,
                "correct_percentage": correct_percentage if test.status == 2 else None,
                "doctor": f"Dr. {test.doctor.first_name} {test.doctor.last_name}"
            })

        return JsonResponse({"tests": test_data})
    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@require_GET
def fetch_test_details(request, test_id):
    percentages = []
    notEmpty = False

    tests = Test.objects.all()

    for test in tests: # calculate correct response %
        percentages.append(_calculate_accuracies(test))

    zippedTests = zip(tests, percentages) # can use both in django for each:  {% for test, percentage in tests %}

    if(tests): # allows webpage to display no tests screen rather than empty table
        notEmpty = True

    return render(request, "testInfo.html", {
        'notEmpty' : notEmpty,
        'tests' : zippedTests
    })

@require_GET
def doctor_tests(request, doctor_id):
    try:
        doctor = Doctor.objects.get(id=doctor_id)
        tests = Test.objects.filter(doctor=doctor)
        
        # Map status codes to human-readable strings
        status_mapping = {
            0: "Not Started",
            1: "In Progress",
            2: "Completed"
        }
        
        test_details = []
        for test in tests:
            stimuli_responses = Stimuli_Response.objects.filter(test=test).select_related('given').values(
                "enum_type", 
                "response", 
                "response_per_click", 
                "given__given_stimuli", 
                "given__correct_order"
            )
            
            test_details.append({
                "test_id": test.test_id,
                "time_started": test.time_started.strftime('%m/%d/%Y, %H:%M:%S') if test.time_started else None,
                "time_ended": test.time_ended.strftime('%m/%d/%Y, %H:%M:%S') if test.time_ended else None,
                "status": status_mapping.get(test.status, "Unknown"),  # Map status to human-readable string
                "patient_age": test.patient_age,
                "stimuli_responses": list(stimuli_responses)
            })
        
        return JsonResponse({
            "doctor": f"{doctor.first_name} {doctor.last_name}",
            "tests": test_details
        })
    except Doctor.DoesNotExist:
        return JsonResponse({"error": "Doctor not found"}, status=404)
    
    
@csrf_exempt
def add_doctor(request):
    """Adds a new doctor, ensuring a unique username in 'doctor#' format."""
    try:
        data = json.loads(request.body)
        first_name = data.get("first_name")
        last_name = data.get("last_name")
        email = data.get("email")

        if not (first_name and last_name and email):
            return JsonResponse({"error": "Missing fields"}, status=400)

        doctor_count = 0
        while True:
            username = f"doctor{doctor_count}"
            if not Doctor.objects.filter(username=username).exists():
                break
            doctor_count += 1

        new_doctor = Doctor.objects.create(username=username, first_name=first_name, last_name=last_name, email=email)
        default_password = "defaultpassword123"
        new_doctor.set_password(default_password)
        new_doctor.save()
 
         # Prepare email content
        subject = "Your Doctor Account Credentials"
        message = (
             f"Hello {first_name},\n\n"
             f"Your account has been created successfully.\n\n"
             f"Username: {username}\n"
             f"Password: {default_password}\n\n"
             f"Please change your password after logging in."
         )
 
         # Attempt to send email
        try:
             send_mail(subject, message, settings.EMAIL_HOST_USER, [email])
             email_status = "Email sent successfully!"
        except Exception:
             email_status = "Email doesn't exist."
 
        return JsonResponse({
             "message": f"Doctor added successfully! {email_status}",
             "doctor_id": new_doctor.id,
             "username": username
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@require_GET
def admin_average_statistics(request):
    try:
        # Call the shared logic for average statistics
        tests = Test.objects.values_list('patient_age', flat=True)

        if not tests:
            age_data = { "30-39": 0, "40-49": 0, "50-59": 0, "60-69": 0,
                         "70-79": 0, "80-89": 0, "90-99": 0, "100+": 0 }
            accuracy_data = {key: 0 for key in age_data.keys()}
            stimulus_accuracy_avg = {}
        else:
            df = pd.DataFrame({'patient_age': tests})
            bins = [30, 39, 49, 59, 69, 79, 89, 99, float('inf')]
            labels = ["30-39", "40-49", "50-59", "60-69", 
                      "70-79", "80-89", "90-99", "100+"]
            df['age_group'] = pd.cut(df['patient_age'], bins=bins, labels=labels, right=True)

            age_data = df['age_group'].value_counts().sort_index().to_dict()

            accuracy_data = {key: [] for key in labels}  

            stimulus_accuracy = {stimulus.id: {label: [] for label in labels} 
                                for stimulus in Given_Stimuli.objects.all()}

            responses = Stimuli_Response.objects.select_related('given')

            for response in responses:
                correct = response.given.correct_order.strip()
                user_response = response.response.strip()
                stimulus_id = response.given.id

                match_count = sum(1 for c1, c2 in zip(correct, user_response) if c1 == c2)
                accuracy_percentage = (match_count / len(correct)) * 100

                age = response.test.patient_age
                age_group = pd.cut([age], bins=bins, labels=labels, right=True)[0]

                if stimulus_id in stimulus_accuracy and age_group in stimulus_accuracy[stimulus_id]:
                    stimulus_accuracy[stimulus_id][age_group].append(accuracy_percentage)

                if age_group in accuracy_data:
                    accuracy_data[age_group].append(accuracy_percentage)

            stimulus_accuracy_avg = {
                stim_id: {
                    age_group: (
                        sum(values) / len(values) if values else 0
                    )
                    for age_group, values in age_dict.items()
                }
                for stim_id, age_dict in stimulus_accuracy.items()
            }

            for key in accuracy_data.keys():
                if accuracy_data[key]:
                    accuracy_data[key] = sum(accuracy_data[key]) / len(accuracy_data[key])
                else:
                    accuracy_data[key] = 0

        # Return the data as JSON
        return JsonResponse({
            'labels': list(age_data.keys()),
            'values': list(age_data.values()),
            'accuracy_labels': list(accuracy_data.keys()),
            'accuracy_values': list(accuracy_data.values()),
            'stimulus_accuracy': stimulus_accuracy_avg
        })
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
    
@require_GET
def overall_average_statistics(request):
     try:
         # Fetch all tests (or apply any filter if needed)
         tests = Test.objects.all()
 
         if not tests:
             return render(request, "admin.html", {
                 "total_tests": 0,
                 "average_patient_age": 0,
                 "average_accuracy": 0,
             })
 
         total_tests = tests.count()
 
         # Calculate overall average patient age
         average_patient_age = tests.aggregate(Avg('patient_age'))['patient_age__avg']
 
         # Calculate overall average accuracy (example based on response data)
         total_accuracy = 0
         total_responses = 0
 
         # Loop through tests to accumulate accuracy data
         for test in tests:
             stimuli_responses = Stimuli_Response.objects.filter(test=test)
             for response in stimuli_responses:
                 total_responses += 1
                 # Compare the response with given stimuli or correct order (whichever is applicable)
                 if response.response == response.given.given_stimuli:  # Assuming given_stimuli holds the correct answer
                     total_accuracy += 1
 
         # Calculate overall accuracy (percentage)
         average_accuracy = (total_accuracy / total_responses * 100) if total_responses > 0 else 0
 
         # Render the admin.html template with the calculated data
         return render(request, "admin.html", {
             "total_tests": total_tests,
             "average_patient_age": average_patient_age,
             "average_accuracy": f"{average_accuracy:.2f}",
         })
 
     except Exception as e:
         return render(request, "admin.html", {
             "error": f"An error occurred: {str(e)}",
             "total_tests": 0,
             "average_patient_age": 0,
             "average_accuracy": 0,
         })

def doctorHomePage(request):
    doctor = request.user

    notifications = Notification.objects.filter(users=doctor)
    
    return render(request, 'doctorHomePage.html', {'notifications': notifications})

# AVG STATS
AGE_BINS = [30, 39, 49, 59, 69, 79, 89, 99, float('inf')]
AGE_LABELS = ["30-39", "40-49", "50-59", "60-69", "70-79", "80-89", "90-99", "100+"]

def get_age_distribution(tests):
    df = pd.DataFrame({'patient_age': [test.patient_age for test in tests]})
    df['age_group'] = pd.cut(df['patient_age'], bins=AGE_BINS, labels=AGE_LABELS, right=True)
    return df['age_group'].value_counts().sort_index().to_dict()

def get_overall_accuracy_by_age(tests):
    from collections import defaultdict

    grouped_accuracies = defaultdict(list)

    for test in tests:
        age = test.patient_age
        age_group = pd.cut([age], bins=AGE_BINS, labels=AGE_LABELS, right=True)[0]
        if age_group:
            accuracy = _calculate_accuracies(test)
            grouped_accuracies[age_group].append(accuracy)
    averaged = {}
    for group in AGE_LABELS:
        values = grouped_accuracies.get(group, [])
        averaged[group] = round(sum(values) / len(values), 2) if values else 0

    return averaged

import numpy as np

def get_overall_accuracy_boxplot_data(tests):
    age_group_accuracies = {label: [] for label in AGE_LABELS}

    for test in tests:
        age = test.patient_age
        if age is None:
            continue
        age_group = pd.cut([age], bins=AGE_BINS, labels=AGE_LABELS, right=True)[0]
        accuracy = _calculate_accuracies(test)
        age_group_accuracies[age_group].append(accuracy)

    # Format for Chart.js boxplot plugin
    boxplot_data = {}

    for group, scores in age_group_accuracies.items():
        if scores:
            sorted_scores = sorted(scores)
            boxplot_data[group] = {
                "min": float(np.min(sorted_scores)),
                "q1": float(np.percentile(sorted_scores, 25)),
                "median": float(np.percentile(sorted_scores, 50)),
                "q3": float(np.percentile(sorted_scores, 75)),
                "max": float(np.max(sorted_scores)),
            }
        else:
            boxplot_data[group] = {
                "min": 0, "q1": 0, "median": 0, "q3": 0, "max": 0
            }

    return boxplot_data

from collections import defaultdict
from statistics import mean

def get_stimulus_latency_data():
    latency_data = defaultdict(lambda: defaultdict(list))

    all_responses = Stimuli_Response.objects.select_related('test', 'given')

    for response in all_responses:
        if not response.given:
            continue

        age_group = response.get_age_group()
        stim_id = response.given.id
        avg_latency = response.get_avg_latency()
        latency_data[stim_id][age_group].append(avg_latency)

    averaged_data = {}
    for stimulus, age_data in latency_data.items():
        averaged_data[stimulus] = {
            age_group: round(mean(latencies), 2)
            for age_group, latencies in age_data.items()
            if latencies
        }

    return averaged_data


def get_stimulus_accuracy_data():
    stimulus_accuracy = {
        stimulus.id: {label: [] for label in AGE_LABELS}
        for stimulus in Given_Stimuli.objects.all()
    }

    responses = Stimuli_Response.objects.select_related('given', 'test')

    for response in responses:
        correct = response.given.correct_order.strip()
        user_response = response.response.strip()
        stimulus_id = response.given.id

        match_count = sum(1 for c1, c2 in zip(correct, user_response) if c1 == c2)
        accuracy_percentage = (match_count / len(correct)) * 100

        age = response.test.patient_age
        age_group = pd.cut([age], bins=AGE_BINS, labels=AGE_LABELS, right=True)[0]

        if stimulus_id in stimulus_accuracy and age_group in stimulus_accuracy[stimulus_id]:
            stimulus_accuracy[stimulus_id][age_group].append(accuracy_percentage)

    stimulus_accuracy_avg = {
        stim_id: {
            age_group: (
                sum(values) / len(values) if values else 0
            )
            for age_group, values in age_dict.items()
        }
        for stim_id, age_dict in stimulus_accuracy.items()
    }

    return stimulus_accuracy_avg



def average_statistics(request):
    tests = list(Test.objects.all())
    
    if not tests:
        age_data = {label: 0 for label in AGE_LABELS}
        accuracy_data = {label: 0 for label in AGE_LABELS}
        accuracy_boxplot_data = {}
        stimulus_accuracy_avg = {} ##
        stimulus_latency = {}
    else:
        age_data = get_age_distribution(tests)
        accuracy_data = get_overall_accuracy_by_age(tests) 
        accuracy_boxplot_data = get_overall_accuracy_boxplot_data(tests)
        stimulus_accuracy_avg = get_stimulus_accuracy_data()
        stimulus_latency = get_stimulus_latency_data()

    context = {
        'accuracy_boxplot': json.dumps(accuracy_boxplot_data),
        'labels': json.dumps(list(age_data.keys())),
        'values': json.dumps(list(age_data.values())),
        'accuracy_labels': json.dumps(list(accuracy_data.keys())),
        'accuracy_values': json.dumps(list(accuracy_data.values())),
        'stimulus_accuracy': json.dumps(stimulus_accuracy_avg),
        'stimulus_latency': json.dumps(stimulus_latency),
    }

    return render(request, 'average_statistics.html', context)

def create_notification_test_completed(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)

            # Extract user_id and test_id from the request data
            user_id = data.get('user_id')
            test_id = data.get('test_id')

            if not user_id or not test_id:
                return JsonResponse({'success': False, 'error': 'Invalid request method.'})
            
            try:
                 test = Test.objects.get(test_id=test_id)

            except Test.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Test matching query does not exist.'})

            if not test.time_ended:
                return JsonResponse({'success': False, 'error': 'Test has not been completed yet.'})

            # Determines if null or not
            test_completion = Stimuli_Response.objects.filter(test=test).values("test__time_ended")

            if not test_completion.exists():
                return JsonResponse({'success': False, 'error': 'Test completion time not found.'})

            # Fetches the first test completion time
            test_time = test_completion.first().get("test__time_ended")

            # Notification message
            message = f"Test Id {test.test_id} has been completed at {test_time}."

            # Fetches the user to recieve the notification
            user = User.objects.get(id=user_id)

            # Fetch doctor associated with test
            doctor = test.doctor 

            #Creates the notification object
            notification = Notification.objects.create(message=message)

            notification.users.add(doctor) # add the notification to the users list of notifications

            return JsonResponse({'success': True, 'notification_id': notification.id})

        except Exception as e:
            # error response incase of exception
            return JsonResponse({'success': False, 'error': str(e)})

    #Checks if not POST
    return JsonResponse({'success': False, 'error': 'Invalid request method.'})

def testComplete(request, testId):
    # get current date time
    now = datetime.now()
    current_datetime = now.strftime("%Y-%m-%d %H:%M:%S")

    test_instance = Test.objects.get(test_id=testId)
    
    test_instance.status = 2
    test_instance.time_ended = current_datetime # gives console warning, but formats date correctly

    Test.save(test_instance)

    return render(request, "testComplete.html", {})

def testStart(request, testId):
    now = datetime.now()
    current_datetime = now.strftime("%Y-%m-%d %H:%M:%S")

    test_instance = Test.objects.get(test_id=testId)
    
    test_instance.status = 1
    test_instance.time_started = current_datetime

    Test.save(test_instance)
    
    return render(request, "instructions.html", {
        'testId' : testId
    })
    
def setting(request, testId, language):
    return render(request, "settings.html", {
        'testId' : testId,
        'language' : language
    })

def forgot_password(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        try:
            user = User.objects.get(username=username)
            token = get_random_string(length=6, allowed_chars='0123456789')
            reset_tokens[username] = token
            send_mail(
                'Password Reset Code',
                f'Your password reset code is: {token}',
                settings.EMAIL_HOST_USER,
                [user.email],
                fail_silently=False,
            )
            return redirect('PPST:reset_password_token')
        except User.DoesNotExist:
            messages.error(request, 'Username not found.')
    return render(request, 'forgot_password.html')

def reset_password_token(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        token = request.POST.get('token')
        if reset_tokens.get(username) == token:
            request.session['reset_username'] = username  # Store username in session for the next step
            return redirect('PPST:reset_password')
        else:
            messages.error(request, 'Invalid token or username.')
    return render(request, 'reset_password_token.html')


def reset_password(request):
    if request.method == 'POST':
        username = request.session.get('reset_username')  # Retrieve username from session
        new_password = request.POST.get('new_password')
        if username:
            try:
                user = User.objects.get(username=username)
                user.set_password(new_password)
                user.save()
                del reset_tokens[username]  # Remove the token after successful reset
                request.session.pop('reset_username')  # Clear session data
                messages.success(request, 'Password reset successfully.')
                return redirect('PPST:doctor_login')
            except User.DoesNotExist:
                messages.error(request, 'Invalid username.')
        else:
            messages.error(request, 'Session expired. Please try again.')
    return render(request, 'reset_password.html')